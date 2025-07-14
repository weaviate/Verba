import base64
import json
import io
import csv

from wasabi import msg

from goldenverba.components.document import Document, create_document
from goldenverba.components.interfaces import Reader
from goldenverba.server.types import FileConfig

# Optional imports with error handling
try:
    from pypdf import PdfReader
except ImportError:
    msg.warn("pypdf not installed, PDF functionality will be limited.")
    PdfReader = None

try:
    import spacy
except ImportError:
    msg.warn("spacy not installed, NLP functionality will be limited.")
    spacy = None

try:
    import docx
except ImportError:
    msg.warn("python-docx not installed, DOCX functionality will be limited.")
    docx = None

try:
    import pandas as pd
except ImportError:
    msg.warn("pandas not installed, Excel functionality will be limited.")
    pd = None

try:
    import openpyxl
except ImportError:
    msg.warn("openpyxl not installed, Excel functionality will be limited.")
    openpyxl = None

try:
    import xlrd
except ImportError:
    msg.warn("xlrd not installed, .xls file functionality will be limited.")
    xlrd = None


class BasicReader(Reader):
    """
    The BasicReader reads text, code, PDF, DOCX, CSV, and Excel files.
    """

    def __init__(self):
        super().__init__()
        self.name = "Default"
        self.description = "Ingests text, code, PDF, DOCX, CSV, and Excel files"
        self.requires_library = ["pypdf", "docx", "spacy", "pandas", "openpyxl"]
        self.extension = [
            ".txt",
            ".py",
            ".js",
            ".html",
            ".css",
            ".md",
            ".mdx",
            ".json",
            ".pdf",
            ".docx",
            ".pptx",
            ".xlsx",
            ".xls",
            ".csv",
            ".ts",
            ".tsx",
            ".vue",
            ".svelte",
            ".astro",
            ".php",
            ".rb",
            ".go",
            ".rs",
            ".swift",
            ".kt",
            ".java",
            ".c",
            ".cpp",
            ".h",
            ".hpp",
        ]  # Add supported text extensions

        # Initialize spaCy model if available
        self.nlp = spacy.blank("en") if spacy else None
        if self.nlp:
            self.nlp.add_pipe("sentencizer", config={"punct_chars": None})

    async def load(self, config: dict, fileConfig: FileConfig) -> list[Document]:
        """
        Load and process a file based on its extension.
        """
        msg.info(f"Loading {fileConfig.filename} ({fileConfig.extension.lower()})")

        if fileConfig.extension != "":
            decoded_bytes = base64.b64decode(fileConfig.content)

        try:
            if fileConfig.extension == "":
                file_content = fileConfig.content
            elif fileConfig.extension.lower() == "json":
                return await self.load_json_file(decoded_bytes, fileConfig)
            elif fileConfig.extension.lower() == "pdf":
                file_content = await self.load_pdf_file(decoded_bytes)
            elif fileConfig.extension.lower() == "docx":
                file_content = await self.load_docx_file(decoded_bytes)
            elif fileConfig.extension.lower() == "csv":
                file_content = await self.load_csv_file(decoded_bytes)
            elif fileConfig.extension.lower() in ["xlsx", "xls"]:
                file_content = await self.load_excel_file(
                    decoded_bytes, fileConfig.extension.lower()
                )
            elif fileConfig.extension.lower() in [
                ext.lstrip(".") for ext in self.extension
            ]:
                file_content = await self.load_text_file(decoded_bytes)
            else:
                try:
                    file_content = await self.load_text_file(decoded_bytes)
                except Exception as e:
                    raise ValueError(
                        f"Unsupported file extension: {fileConfig.extension}"
                    )

            return [create_document(file_content, fileConfig)]
        except Exception as e:
            msg.fail(f"Failed to load {fileConfig.filename}: {str(e)}")
            raise

    async def load_text_file(self, decoded_bytes: bytes) -> str:
        """Load and decode a text file."""
        try:
            return decoded_bytes.decode("utf-8")
        except UnicodeDecodeError:
            # Fallback to latin-1 if UTF-8 fails
            return decoded_bytes.decode("latin-1")

    async def load_json_file(
        self, decoded_bytes: bytes, fileConfig: FileConfig
    ) -> list[Document]:
        """Load and parse a JSON file."""
        try:
            json_obj = json.loads(decoded_bytes.decode("utf-8"))
            document = Document.from_json(json_obj, self.nlp)
            return (
                [document]
                if document
                else [create_document(json.dumps(json_obj, indent=2), fileConfig)]
            )
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in {fileConfig.filename}: {str(e)}")

    async def load_pdf_file(self, decoded_bytes: bytes) -> str:
        """Load and extract text from a PDF file."""
        if not PdfReader:
            raise ImportError("pypdf is not installed. Cannot process PDF files.")
        pdf_bytes = io.BytesIO(decoded_bytes)
        reader = PdfReader(pdf_bytes)
        return "\n\n".join(page.extract_text() for page in reader.pages)

    async def load_docx_file(self, decoded_bytes: bytes) -> str:
        """Load and extract text from a DOCX file."""
        if not docx:
            raise ImportError(
                "python-docx is not installed. Cannot process DOCX files."
            )
        docx_bytes = io.BytesIO(decoded_bytes)
        reader = docx.Document(docx_bytes)
        return "\n".join(paragraph.text for paragraph in reader.paragraphs)

    async def load_csv_file(self, decoded_bytes: bytes) -> str:
        """Load and convert CSV file to readable text format."""
        try:
            # Try UTF-8 first, fallback to latin-1
            try:
                text_content = decoded_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text_content = decoded_bytes.decode("latin-1")

            csv_reader = csv.reader(io.StringIO(text_content))
            rows = list(csv_reader)

            if not rows:
                return "Empty CSV file"

            # Format as a readable table
            result = []
            headers = rows[0] if rows else []

            # Add headers
            if headers:
                result.append("Headers: " + " | ".join(headers))
                result.append(" \n\n")

            # Add data rows
            for i, row in enumerate(rows[1:], 1):
                if len(row) == len(headers):
                    row_data = []
                    for header, value in zip(headers, row):
                        row_data.append(f"{header}: {value}")
                    result.append(f"Row {i}: {' | '.join(row_data)}")
                else:
                    # Handle rows with different column counts
                    result.append(f"Row {i}: {' | '.join(row)}")
                result.append(" \n\n")
            return "\n".join(result)

        except Exception as e:
            raise ValueError(f"Error reading CSV file: {str(e)}")

    async def load_excel_file(self, decoded_bytes: bytes, extension: str) -> str:
        """Load and convert Excel file to readable text format."""
        if not pd and not openpyxl:
            raise ImportError("pandas or openpyxl is required to process Excel files.")

        try:
            excel_bytes = io.BytesIO(decoded_bytes)

            # Use pandas if available for better support
            if pd:
                # Read all sheets
                if extension == "xlsx":
                    sheets_dict = pd.read_excel(
                        excel_bytes, sheet_name=None, engine="openpyxl"
                    )
                else:  # xls
                    try:
                        sheets_dict = pd.read_excel(
                            excel_bytes, sheet_name=None, engine="xlrd"
                        )
                    except Exception as e:
                        # Try auto engine detection as fallback
                        try:
                            sheets_dict = pd.read_excel(
                                excel_bytes, sheet_name=None, engine=None
                            )
                        except Exception:
                            raise ImportError(
                                f"Cannot read .xls file. Please install 'xlrd' for .xls support: pip install xlrd. "
                                f"Original error: {str(e)}"
                            )

                result = []

                for sheet_name, df in sheets_dict.items():
                    result.append(f"\nSheet: {sheet_name}")

                    if df.empty:
                        result.append("(Empty sheet)")
                        continue

                    result.append(" \n\n")

                    # Add column headers
                    headers = df.columns.tolist()
                    result.append("Headers: " + " | ".join(str(h) for h in headers))
                    result.append(" \n\n")

                    for idx, (_, row) in enumerate(df.iterrows()):
                        row_data = []
                        for header, value in zip(headers, row):
                            # Handle NaN values
                            display_value = str(value) if pd.notna(value) else ""
                            row_data.append(f"{header}: {display_value}")
                        result.append(f"Row {idx + 1}: {' | '.join(row_data)}")
                        result.append(" \n\n")

                return "\n".join(result)

            else:
                # Fallback to openpyxl for basic reading
                if extension != "xlsx":
                    raise ImportError(
                        "openpyxl only supports .xlsx files. Please install pandas for .xls support."
                    )

                from openpyxl import load_workbook

                workbook = load_workbook(excel_bytes, data_only=True)

                result = []

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    result.append(f"\nSheet: {sheet_name}")
                    result.append(" \n\n")

                    rows_data = []
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):  # Skip empty rows
                            rows_data.append(
                                [str(cell) if cell is not None else "" for cell in row]
                            )

                    if not rows_data:
                        result.append("(Empty sheet)")
                        continue

                    # Add headers and data
                    headers = rows_data[0] if rows_data else []
                    result.append("Headers: " + " | ".join(headers))
                    result.append(" \n\n")

                    for i, row in enumerate(rows_data[1:], 1):
                        if len(row) == len(headers):
                            row_data = [f"{h}: {v}" for h, v in zip(headers, row)]
                            result.append(f"Row {i}: {' | '.join(row_data)}")
                            result.append(" \n\n")
                        else:
                            result.append(f"Row {i}: {' | '.join(row)}")
                            result.append(" \n\n")

                return "\n".join(result)

        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
